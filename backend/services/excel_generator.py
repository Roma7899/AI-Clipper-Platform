import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any

class ExcelGenerator:
    def __init__(self):
        pass

    def generate_results_excel(self, video_info: Dict[str, Any], clips: List[Dict[str, Any]], output_path: str):
        """Generates a professional Excel sheet with clip results."""
        wb = openpyxl.Workbook()
        
        # --- Sheet 1: Clips Results ---
        ws1 = wb.active
        ws1.title = "Clips Results"
        
        headers = [
            "#", "Clip Link", "Viral Score", "Title (EN)", "Title (AR)", 
            "Type", "Hook Line", "Keywords", "Description (EN)", 
            "Description (AR)", "Emojis", "Platform", "Duration", 
            "Start Time", "End Time", "Created At"
        ]
        
        # Styling
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Write data
        for i, clip in enumerate(clips, 1):
            row = i + 1
            ws1.cell(row=row, column=1, value=i)
            
            # Clip Link (Hyperlink)
            link_cell = ws1.cell(row=row, column=2, value="Download Clip")
            link_cell.hyperlink = clip.get("clip_url", "")
            link_cell.font = Font(color="0000FF", underline="single")
            
            # Viral Score with color coding
            score = clip.get("viral_score", 0)
            score_cell = ws1.cell(row=row, column=3, value=score)
            if score >= 85:
                score_cell.fill = PatternFill(start_color="00F5A0", end_color="00F5A0", fill_type="solid")
            elif score >= 70:
                score_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            else:
                score_cell.fill = PatternFill(start_color="FF4C4C", end_color="FF4C4C", fill_type="solid")
            
            ws1.cell(row=row, column=4, value=clip.get("suggested_title", ""))
            ws1.cell(row=row, column=5, value=clip.get("suggested_title_ar", ""))
            ws1.cell(row=row, column=6, value=clip.get("clip_type", ""))
            ws1.cell(row=row, column=7, value=clip.get("hook_line", ""))
            ws1.cell(row=row, column=8, value=", ".join(clip.get("keywords", [])))
            ws1.cell(row=row, column=9, value=clip.get("description", ""))
            ws1.cell(row=row, column=10, value=clip.get("description_ar", ""))
            ws1.cell(row=row, column=11, value=clip.get("emojis", ""))
            ws1.cell(row=row, column=12, value=clip.get("platform_recommendation", ""))
            
            duration = clip.get("end_time", 0) - clip.get("start_time", 0)
            ws1.cell(row=row, column=13, value=f"{duration:.2f}s")
            ws1.cell(row=row, column=14, value=clip.get("start_time", 0))
            ws1.cell(row=row, column=15, value=clip.get("end_time", 0))
            ws1.cell(row=row, column=16, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

            # Alternate row coloring
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    if col != 3: # Don't overwrite viral score color
                        ws1.cell(row=row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Column widths and wrap text
        for col in range(1, len(headers) + 1):
            ws1.column_dimensions[get_column_letter(col)].width = 20
            for row in range(1, len(clips) + 2):
                ws1.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="center")

        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = ws1.dimensions

        # --- Sheet 2: Video Info ---
        ws2 = wb.create_sheet("Video Info")
        ws2.append(["Field", "Value"])
        ws2.append(["Original Title", video_info.get("title", "")])
        ws2.append(["Original URL", video_info.get("original_url", "Uploaded File")])
        ws2.append(["Total Clips Found", len(clips)])
        ws2.append(["Video Summary (EN)", video_info.get("video_summary", "")])
        ws2.append(["Video Summary (AR)", video_info.get("video_summary_ar", "")])
        ws2.append(["Processed At", datetime.now().strftime("%Y-%m-%d %H:%M")])
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 80

        wb.save(output_path)
